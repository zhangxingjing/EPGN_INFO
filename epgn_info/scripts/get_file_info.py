# coding: utf-8
import os
import re
import shutil

import pymysql
from openpyxl import load_workbook


class ParseFile():
    """传入以及文件路径，获取当前文件中所有hdf文件位置，插入数据库"""

    def __init__(self):
        self.db = pymysql.connect(host='127.0.0.1', user='root', passwd="root", db='epgn')  # 连接数据库
        self.main_path = '/home/small-spider/Desktop'

    def get_def_file(self, main_path):
        dir_paths = []
        xlsm_paths = []
        dir_file = os.listdir(main_path)
        for file_name in dir_file:
            # 这里判断文件夹的时候， 必须加上当前文件的路径
            file_path = main_path + '/' + file_name
            if os.path.isdir(file_path):
                # 如果当前是目录，进去遍历文件名
                dir_2_file = os.listdir(file_path)

                for file_2_name in dir_2_file:
                    # 这是"2018.06.29"中的文件名
                    file_2_path = main_path + '/' + file_name + '/' + file_2_name
                    if ".xlsm" in file_2_path:
                        xlsm_paths.append(file_2_path)
                    else:
                        dir_paths.append(file_2_path)
        return xlsm_paths, dir_paths

    def parse_xlsm(self, xlsm_path):
        wb = load_workbook(xlsm_path)
        sheet = wb.get_sheet_by_name("Sheet1")

        # 获取指定单元格的值
        car_num = sheet.cell(4, 2).value
        car_model = sheet.cell(5, 2).value
        propulsion = sheet.cell(7, 2).value
        power = sheet.cell(8, 2).value
        parts = sheet.cell(18, 2).value
        working = sheet.cell(22, 3).value

        return car_num, car_model, propulsion, power, parts, working

    def parse_hdf(self, dir_path):
        hdf_list = []
        # for dir_path in dir_path_list:
        dir_name = os.listdir(dir_path)
        for file_name in dir_name:
            if ".hdf" in file_name:
                # 这里才是需要的文件地址 /home/small-spider/Desktop/2018.06.29 BSUV BV6-407/A18/F3_VZ_run10.hdf
                hdf_file_path = dir_path + "/" + file_name
                hdf_list.append(hdf_file_path)
        return hdf_list

    def get_item(self):
        xlsm_path_list, dir_paths_list = self.get_def_file(self.main_path)
        # items = []
        for xlsm_path in xlsm_path_list:
            car_num, car_model, propulsion, power, parts, working = self.parse_xlsm(xlsm_path)
            # print(car_num, car_model, propulsion, power, parts, working)
        for dir_path in dir_paths_list:
            # print(dir_path)
            hdf_file_path_list = self.parse_hdf(dir_path)
            print(len(hdf_file_path_list))  # 19 /49 /47 /1 /56 ==> 172
            for hdf_file_path in hdf_file_path_list:
                item = {
                    "车号": car_num,
                    "车型": car_model,
                    "动力总成": propulsion,
                    "功率": power,
                    "零部件": parts,
                    "工况": working,
                    "文件位置": hdf_file_path
                }
                print(item)
                # items.append(item)  # 172
                yield item

    def parse_sql(self):
        cur = self.db.cursor()  # 创建游标
        num = 1
        for item in self.get_item():
            # item["num"] = num
            try:
                sql_info = """insert into tb_car values (%d, \"%s\", \"%s\", \"%s\", \"%s\", \"%s\", \"%s\", \"%s\")""" % (
                    num, item["车号"], item["车型"], item["动力总成"], item["功率"], item["零部件"], item["工况"], item["文件位置"])
                cur.execute(sql_info)  # 执行SQL语句
                self.db.commit()
                # print(item)  # 拿到完整的数据，在这里把数据插入数据库
            except Exception as e:
                # print("数据插入失败...")
                # print(e)
                self.db.rollback()  # 使用数据库连接对象回滚
                continue
            num += 1
            # break


class CopyFile():
    """传入文件夹绝对路径, 把文件复制到制定的路径下"""

    def __init__(self):
        self.path = "/home/small-spider/Music/"

    def copy_file(self):
        """
        1. 获取文件名
        2. 修改文件名
        3. 复制文件
        ==> 在前端页面里面获取的只是文件名？？？ ==> 怎么获取完整路径
        """
        file_info = "/home/small-spider/Desktop/2018.04.29 NMS PHEV CVFHAI NAN B8P/20180429 NMS PHEV B8P-508 A06/BCM D mit AC run01.hdf"  # 文件的原始路径

        # 通过原始路径, 获取文件名
        file_path_1 = re.findall(r"(.*)/(.*)", file_info)[0][0] + "/"  # 文件所在的位置(文件夹)
        file_name = re.findall(r"(.*)/(.*)", file_info)[0][1]
        # print(file_path_1, "\n", file_name)

        # 先复制文件, 再修改文件名
        shutil.copy(file_info, self.path)

        new_file_name = "2019_05_21.hdf"
        os.rename(self.path + file_name, self.path + new_file_name)


class FilePath():
    def get_file_path(self):
        path_1 = "/media/sf_E_DRIVE/FileInfo"
        path_2 = "E:\\MQB A0\\AGA&ANS"



        print(path_1 + '/', path_2 + "\\")



class GetFile():
    # 前端上传文件的时候，上传的是整个文件吗，还是说只是文件名
    # 怎么获取前端上传文件的所有内容
    # 把上传文件的内容通过网络传送给服务器
    def get_file_content(self):
        with open('/media/small-spider/文件/大众/file/qwe_run01.asc', 'r') as f1:
            list1 = f1.readlines()
            print(list1)